# What you get in memory

# A list called tables. One entry per TABLE block found by Textract. Each entry is a dict with:
# page: Page number the table is on.
# confidence: Textract’s confidence for the table block (if present).
# bbox and polygon: The table’s bounding box and polygon (normalized 0–1 coordinates).
# matrix: A 2D list of strings representing the table as rows and columns.
# Merged cells: the text is placed in the top-left anchor cell; the other merged positions are left empty.
# cells: A list of cell dicts with rich metadata:
# row_index, column_index: 1-based indices.
# row_span, column_span: spans for merged cells.
# text: concatenated WORDs (and [x]/[ ] for checkboxes).
# confidence: Textract’s confidence for the cell.
# bbox, polygon: cell geometry (normalized 0–1).
# What gets written to disk

# One CSV per table using matrix only:
# File name: page_{page}table{i}.csv in the out_dir you pass.
# Each CSV row equals one table row; each column equals one table column.
# Quick example of what tables[0] might look like

# tables[0]["page"] -> 1
# tables[0]["matrix"] -> [
# ["Date of Service", "CPT", "Units", "Charge", "Patient Resp"],
# ["01/12/2024", "99213", "1", "
# 150.00
# "
# ,
# "
# 150.00","30.00"],
# ...
# ]
# tables[0]["cells"][0] -> {
# "row_index": 1,
# "column_index": 1,
# "row_span": 1,
# "column_span": 1,
# "text": "Date of Service",
# "confidence": 98.3,
# "bbox": {"Left": 0.12, "Top": 0.18, "Width": 0.15, "Height": 0.02},
# "polygon": [...]
# }
# Notes

# Coordinates (bbox/polygon) are relative to the page (0–1). Convert to pixels/points by multiplying by page width/height if needed.
# Checkboxes in cells appear as “[x]” or “[ ]”.
# If Textract finds no tables, tables = [] and no CSVs are written.
# The CSVs only include text; if you need geometry/confidence exported too, you can dump tables to a JSON file or build a custom CSV writer for the cells list.
# How to see it quickly

# After running:
# print(len(tables)) # number of tables
# print(tables[0]["page"], len(tables[0]["matrix"]), len(tables[0]["matrix"][0])) # page, row count, column count
# print(tables[0]["cells"][0]) # first cell’s metadata and text

import json
from typing import List, Dict, Any, Tuple, Optional
import csv
import os

def load_textract_json(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def build_block_map(blocks: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    return {b["Id"]: b for b in blocks}

def get_text_from_block(block: Dict[str, Any], block_map: Dict[str, Dict[str, Any]]) -> str:
    # Concatenate WORDs and handle checkboxes (SELECTION_ELEMENT)
    texts = []
    for rel in block.get("Relationships", []):
        if rel["Type"] == "CHILD":
            for cid in rel["Ids"]:
                child = block_map.get(cid)
                if not child:
                    continue
                if child["BlockType"] == "WORD":
                    texts.append(child["Text"])
                elif child["BlockType"] == "SELECTION_ELEMENT":
                    status = child.get("SelectionStatus", "NOT_SELECTED")
                    texts.append("[x]" if status == "SELECTED" else "[ ]")
    return " ".join(texts).strip()

def extract_table_cells(
    table_block: Dict[str, Any],
    block_map: Dict[str, Dict[str, Any]]
) -> List[Dict[str, Any]]:
    cell_ids = []
    for rel in table_block.get("Relationships", []):
        if rel["Type"] == "CHILD":
            cell_ids.extend(rel["Ids"])
    cells = []
    for cid in cell_ids:
        cell = block_map.get(cid)
        if not cell or cell["BlockType"] not in ("CELL", "MERGED_CELL"):
            continue
        cell_text = get_text_from_block(cell, block_map)
        cells.append({
            "row_index": cell.get("RowIndex", 1),
            "column_index": cell.get("ColumnIndex", 1),
            "row_span": cell.get("RowSpan", 1),
            "column_span": cell.get("ColumnSpan", 1),
            "text": cell_text,
            "confidence": cell.get("Confidence"),
            "bbox": cell.get("Geometry", {}).get("BoundingBox"),
            "polygon": cell.get("Geometry", {}).get("Polygon"),
        })
    return cells

def build_table_matrix(cells: List[Dict[str, Any]]) -> List[List[str]]:
    if not cells:
        return []
    max_row = 0
    max_col = 0
    for c in cells:
        max_row = max(max_row, c["row_index"] + c["row_span"] - 1)
        max_col = max(max_col, c["column_index"] + c["column_span"] - 1)
    matrix = [["" for _ in range(max_col)] for _ in range(max_row)]
    # Place text; for merged cells, put text in the top-left anchored cell
    for c in cells:
        r0 = c["row_index"] - 1
        c0 = c["column_index"] - 1
        for dr in range(c["row_span"]):
            for dc in range(c["column_span"]):
                rr = r0 + dr
                cc = c0 + dc
                if dr == 0 and dc == 0:
                    matrix[rr][cc] = c["text"]
                else:
                    # Optional: mark merged positions or leave empty
                    if not matrix[rr][cc]:
                        matrix[rr][cc] = ""
    return matrix

def extract_tables(textract_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    blocks = textract_json.get("Blocks", [])
    block_map = build_block_map(blocks)
    tables = []
    for b in blocks:
        if b["BlockType"] == "TABLE":
            cells = extract_table_cells(b, block_map)
            matrix = build_table_matrix(cells)
            tables.append({
                "page": b.get("Page"),
                "confidence": b.get("Confidence"),
                "bbox": b.get("Geometry", {}).get("BoundingBox"),
                "polygon": b.get("Geometry", {}).get("Polygon"),
                "cells": cells,
                "matrix": matrix,
            })
    return tables

def save_tables_to_csv(tables: List[Dict[str, Any]], out_dir: str) -> None:
    os.makedirs(out_dir, exist_ok=True)
    for i, t in enumerate(tables, start=1):
        page = t.get("page", "unknown")
        path = os.path.join(out_dir, f"page_{page}_table_{i}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in t["matrix"]:
                writer.writerow(row)

# Example usage:
# data = load_textract_json("textract_output.json")
# tables = extract_tables(data)
# save_tables_to_csv(tables, "./tables_out")
# Now you also have rich metadata in tables[i]["cells"].

# Great question. You can detect merged cells in Textract tables by checking either:

# RowSpan > 1 or ColumnSpan > 1 on CELL blocks, and/or
# BlockType == "MERGED_CELL" (when present).
# Below are small additions to your current code to:

# Capture the block type for each cell.
# Compute which grid positions a merged cell covers.
# Extract only merged-cell info per table.
# Optionally detect header-merged cells (likely in the top of the table).
# Additions/changes to your code

# Keep the block type and the block Id for each cell:
# Update extract_table_cells to include block_type and id.
# Utility to list merged cells and the positions they cover:
# A merged cell is one where row_span > 1 or column_span > 1, or block_type == "MERGED_CELL".
# Optional: flag merged cells that are likely headers:
# Consider cells in the first table row(s) or within the top X% of the table’s bounding box.
# Code snippets (drop-in additions)

# Replace your extract_table_cells with this version to include block_type and id:

def extract_table_cells(
    table_block: Dict[str, Any],
    block_map: Dict[str, Dict[str, Any]]
) -> List[Dict[str, Any]]:
    cell_ids = []
    for rel in table_block.get("Relationships", []):
        if rel["Type"] == "CHILD":
            cell_ids.extend(rel["Ids"])
    cells = []
    for cid in cell_ids:
        cell = block_map.get(cid)
        if not cell or cell["BlockType"] not in ("CELL", "MERGED_CELL"):
            continue
        cell_text = get_text_from_block(cell, block_map)
        cells.append({
            "id": cell.get("Id"),
            "block_type": cell.get("BlockType"),
            "row_index": cell.get("RowIndex", 1),
            "column_index": cell.get("ColumnIndex", 1),
            "row_span": cell.get("RowSpan", 1),
            "column_span": cell.get("ColumnSpan", 1),
            "text": cell_text,
            "confidence": cell.get("Confidence"),
            "bbox": cell.get("Geometry", {}).get("BoundingBox"),
            "polygon": cell.get("Geometry", {}).get("Polygon"),
        })
    return cells

# Add helpers to find merged cells and compute covered positions:

def covered_positions(cell: Dict[str, Any]) -> List[Tuple[int, int]]:
    # Returns 1-based (row, col) positions covered by this cell’s span
    positions = []
    for r in range(cell["row_index"], cell["row_index"] + cell["row_span"]):
        for c in range(cell["column_index"], cell["column_index"] + cell["column_span"]):
            positions.append((r, c))
    return positions

def get_merged_cells_from_table(table: Dict[str, Any]) -> List[Dict[str, Any]]:
    merged = []
    for c in table["cells"]:
        is_merged = (c["row_span"] > 1 or c["column_span"] > 1 or c.get("block_type") == "MERGED_CELL")
        if is_merged:
            mc = dict(c)
            mc["covers"] = covered_positions(c)  # list of (row, col)
            merged.append(mc)
    return merged

# Optional: identify merged header cells

# Heuristics: merged cells in the first row(s) or within the top 15–20% of the table height.

def is_probable_header_merged_cell(merged_cell: Dict[str, Any], table: Dict[str, Any], top_fraction: float = 0.2) -> bool:
    # Row-based heuristic
    if merged_cell["row_index"] <= 2:
        return True
    # Geometry-based heuristic (uses normalized coordinates)
    table_bbox = table.get("bbox") or {}
    cell_bbox = merged_cell.get("bbox") or {}
    if not table_bbox or not cell_bbox:
        return False
    if table_bbox.get("Height", 0) <= 0:
        return False
    rel_top = (cell_bbox["Top"] - table_bbox["Top"]) / table_bbox["Height"]
    return rel_top <= top_fraction

# exmaple usage
data = load_textract_json("textract_output.json")
tables = extract_tables(data)

for ti, t in enumerate(tables, start=1):
    merged_cells = get_merged_cells_from_table(t)
    print(f"Table #{ti} on page {t.get('page')}: {len(merged_cells)} merged cell(s)")

    # Print merged cell info
    for m in merged_cells:
        print(
            f"  Anchor (r{m['row_index']}, c{m['column_index']}), "
            f"span r{m['row_span']} x c{m['column_span']}, "
            f"text='{m['text']}', covers={m['covers']}"
        )

    # If you only want header-merged cells:
    header_merged = [m for m in merged_cells if is_probable_header_merged_cell(m, t)]
    print(f"  Header-like merged cells: {len(header_merged)}")

# What you’ll get for a typical header merge

# For a header like “Service Details” that spans columns 1–3 on the first row:
# row_index=1, column_index=1, row_span=1, column_span=3
# covers=[(1,1), (1,2), (1,3)]
# text="Service Details"
# block_type might be "CELL" or "MERGED_CELL" depending on what Textract emitted.
# Notes and tips

# Textract does not explicitly label “header” cells; you must infer them with row index or geometry.
# Some outputs include MERGED_CELL blocks without text; the text is typically associated with the CELL at the top-left anchor. The RowSpan/ColumnSpan approach is the most reliable.
# If you need only merged cells across all tables, just collect get_merged_cells_from_table(t) for each t and ignore t["matrix"].
# For auditing, you can export merged cell metadata to a CSV: page, table_index, anchor_row, anchor_col, row_span, col_span, text, confidence.

Add these helpers
def covered_positions(cell):
positions = []
for r in range(cell["row_index"], cell["row_index"] + cell["row_span"]):
for c in range(cell["column_index"], cell["column_index"] + cell["column_span"]):
positions.append((r, c))
return positions
def get_merged_cells_from_table(table):
merged = []
for c in table["cells"]:
if c.get("row_span", 1) > 1 or c.get("column_span", 1) > 1 or c.get("block_type") == "MERGED_CELL":
mc = dict(c)
mc["covers"] = covered_positions(c)
merged.append(mc)
return merged

def export_merged_cells_csv(tables, out_path):
from collections import defaultdict
page_counts = defaultdict(int)
with open(out_path, "w", newline="", encoding="utf-8") as f:
writer = csv.writer(f)
writer.writerow([
"page", "table_global_index", "table_index_on_page",
"anchor_row", "anchor_col", "row_span", "col_span",
"text", "confidence", "covers"
])
for gi, t in enumerate(tables, start=1):
page = t.get("page")
page_counts[page] += 1
on_page_idx = page_counts[page]
merged_cells = get_merged_cells_from_table(t)
for m in merged_cells:
covers_str = ";".join([f"{r}:{c}" for (r, c) in m["covers"]])
writer.writerow([
page, gi, on_page_idx,
m.get("row_index"), m.get("column_index"),
m.get("row_span", 1), m.get("column_span", 1),
m.get("text", ""), m.get("confidence", ""),
covers_str
])

# Nice work. Two key points now:

# CSV can’t preserve merged cells. You won’t be able to detect merges from CSV alone.
# You already have merge info in the parsed tables (row_span/column_span, and sometimes BlockType == MERGED_CELL). Use that to export a “merged-cells summary” or write an Excel file with real merged ranges.
# Below are drop-in helpers to:

# Produce a merged-cells-only CSV across all tables.
# Write an Excel workbook that visually preserves merges.
# If your current cell dicts don’t include block_type/id, they’re optional; merges are determined by row_span/column_span.

# Export a merged-cells summary CSV
# Each row = one merged cell (anchor position + span + text) for every table.
# Requirements: none (uses Python stdlib).

merged_csv_path = "merged_cells_summary.csv"
export_merged_cells_csv(tables, merged_csv_path)

# Write an Excel workbook with real merged ranges
# This preserves the visual layout of merged headers and cells.
# Requirements:

# pip install openpyxl
# Code:
def save_tables_to_excel_with_merges(tables, out_xlsx):
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
# remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

page_counts = defaultdict(int)

for gi, t in enumerate(tables, start=1):
    page = t.get("page")
    page_counts[page] += 1
    on_page_idx = page_counts[page]

    # Sheet name max 31 chars
    sheet_name = f"p{page}_t{on_page_idx}"
    sheet_name = sheet_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Write matrix text first
    matrix = t.get("matrix", [])
    for row in matrix:
        ws.append(row)

    # Apply merged ranges
    merged_cells = get_merged_cells_from_table(t)
    for m in merged_cells:
        r1 = int(m["row_index"])
        c1 = int(m["column_index"])
        r2 = r1 + int(m.get("row_span", 1)) - 1
        c2 = c1 + int(m.get("column_span", 1)) - 1
        if r2 > r1 or c2 > c1:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            # Optional: center + wrap
            cell = ws.cell(row=r1, column=c1)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

wb.save(out_xlsx)

save_tables_to_excel_with_merges(tables, "textract_tables_with_merges.xlsx")

# How to use with what you have now

# If you still have the tables variable in memory from parsing Textract JSON: call export_merged_cells_csv and/or save_tables_to_excel_with_merges directly.
# If you only saved CSVs and discarded the tables variable, you cannot recover merges from CSVs. Re-run your parser on the same Textract JSON to get tables again, then export merges.
# Optional: detect header-merged cells only

# Simple heuristic: merged cells in the first 1–2 rows of the table.
def merged_header_cells_only(table, max_header_rows=2):
merged = get_merged_cells_from_table(table)
return [m for m in merged if m["row_index"] <= max_header_rows]

# Below is a complete, ready-to-run toolkit that will:

# Parse Textract JSON and extract all tables across pages.
# Capture full cell metadata (row/col, spans, confidence, geometry).
# Identify all merged cells (including header merges).
# Export per-table CSVs (text only), a merged-cells summary CSV, a JSON metadata dump, and an Excel workbook that visually preserves merged ranges.
# What you’ll need

# Python 3.8+ (3.10+ recommended)
# pip install openpyxl
# Your Textract JSON output file(s)
# Single-file script (save as textract_tables_tool.py)

# Supports a single JSON file or a directory of multiple Textract JSONs (e.g., async job pages). It merges blocks from all files automatically.

import argparse
import csv
import json
import os
from typing import List, Dict, Any, Tuple

# =========================
# Loading and parsing JSON
# =========================

def load_textract_json_any(path: str) -> Dict[str, Any]:
    """
    Load a Textract JSON. If path is a directory, merge all *.json files (concatenate Blocks).
    Returns a dict with a single 'Blocks' list, which works for both sync and async outputs.
    """
    if os.path.isdir(path):
        blocks = []
        for fname in sorted(os.listdir(path)):
            if not fname.lower().endswith(".json"):
                continue
            fpath = os.path.join(path, fname)
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
                blocks.extend(data.get("Blocks", []))
        return {"Blocks": blocks}
    else:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        # Normalize to a dict with Blocks
        if isinstance(data, dict) and "Blocks" in data:
            return data
        elif isinstance(data, list):
            # Some people store blocks directly as a list
            return {"Blocks": data}
        else:
            raise ValueError("Input JSON does not look like a Textract response (missing 'Blocks').")

def build_block_map(blocks: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    return {b.get("Id"): b for b in blocks if "Id" in b}

def get_text_from_block(block: Dict[str, Any], block_map: Dict[str, Dict[str, Any]]) -> str:
    texts = []
    for rel in block.get("Relationships", []):
        if rel.get("Type") == "CHILD":
            for cid in rel.get("Ids", []):
                child = block_map.get(cid)
                if not child:
                    continue
                bt = child.get("BlockType")
                if bt == "WORD":
                    texts.append(child.get("Text", ""))
                elif bt == "SELECTION_ELEMENT":
                    status = child.get("SelectionStatus", "NOT_SELECTED")
                    texts.append("[x]" if status == "SELECTED" else "[ ]")
    return " ".join(t for t in texts if t).strip()

def extract_table_cells(table_block: Dict[str, Any], block_map: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    cell_ids = []
    for rel in table_block.get("Relationships", []):
        if rel.get("Type") == "CHILD":
            cell_ids.extend(rel.get("Ids", []))
    cells = []
    for cid in cell_ids:
        cell = block_map.get(cid)
        if not cell or cell.get("BlockType") not in ("CELL", "MERGED_CELL"):
            continue
        cell_text = get_text_from_block(cell, block_map)
        cells.append({
            "id": cell.get("Id"),
            "block_type": cell.get("BlockType"),
            "row_index": int(cell.get("RowIndex", 1)),
            "column_index": int(cell.get("ColumnIndex", 1)),
            "row_span": int(cell.get("RowSpan", 1)),
            "column_span": int(cell.get("ColumnSpan", 1)),
            "text": cell_text,
            "confidence": cell.get("Confidence"),
            "bbox": (cell.get("Geometry", {}) or {}).get("BoundingBox"),
            "polygon": (cell.get("Geometry", {}) or {}).get("Polygon"),
        })
    return cells

def build_table_matrix(cells: List[Dict[str, Any]]) -> List[List[str]]:
    if not cells:
        return []
    max_row = 0
    max_col = 0
    for c in cells:
        max_row = max(max_row, c["row_index"] + c["row_span"] - 1)
        max_col = max(max_col, c["column_index"] + c["column_span"] - 1)
    matrix = [["" for _ in range(max_col)] for _ in range(max_row)]
    # place text; merged text goes to top-left anchor cell
    for c in cells:
        r0 = c["row_index"] - 1
        c0 = c["column_index"] - 1
        for dr in range(c["row_span"]):
            for dc in range(c["column_span"]):
                rr = r0 + dr
                cc = c0 + dc
                if dr == 0 and dc == 0:
                    matrix[rr][cc] = c["text"]
                else:
                    if not matrix[rr][cc]:
                        matrix[rr][cc] = ""
    return matrix

def extract_tables(textract_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    blocks = textract_json.get("Blocks", [])
    block_map = build_block_map(blocks)
    tables = []
    for b in blocks:
        if b.get("BlockType") == "TABLE":
            cells = extract_table_cells(b, block_map)
            matrix = build_table_matrix(cells)
            tables.append({
                "page": b.get("Page"),
                "confidence": b.get("Confidence"),
                "bbox": (b.get("Geometry", {}) or {}).get("BoundingBox"),
                "polygon": (b.get("Geometry", {}) or {}).get("Polygon"),
                "cells": cells,
                "matrix": matrix,
            })
    annotate_table_indices(tables)
    return tables

def annotate_table_indices(tables: List[Dict[str, Any]]) -> None:
    # Adds table indices to each table dict: global_index and index_on_page
    page_counts = {}
    for i, t in enumerate(tables, start=1):
        t["global_index"] = i
        page = t.get("page")
        page_counts[page] = page_counts.get(page, 0) + 1
        t["index_on_page"] = page_counts[page]

# =========================
# Merged cell utilities
# =========================

def covered_positions(cell: Dict[str, Any]) -> List[Tuple[int, int]]:
    pos = []
    for r in range(cell["row_index"], cell["row_index"] + cell["row_span"]):
        for c in range(cell["column_index"], cell["column_index"] + cell["column_span"]):
            pos.append((r, c))
    return pos

def get_merged_cells_from_table(table: Dict[str, Any]) -> List[Dict[str, Any]]:
    merged = []
    for c in table.get("cells", []):
        if c.get("row_span", 1) > 1 or c.get("column_span", 1) > 1 or c.get("block_type") == "MERGED_CELL":
            mc = dict(c)
            mc["covers"] = covered_positions(c)
            merged.append(mc)
    return merged

def merged_header_cells_only(table: Dict[str, Any], max_header_rows: int = 2) -> List[Dict[str, Any]]:
    merged = get_merged_cells_from_table(table)
    return [m for m in merged if m["row_index"] <= max_header_rows]

# =========================
# Exporters
# =========================

def ensure_outdir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def save_tables_to_csv(tables: List[Dict[str, Any]], out_dir: str, prefix: str = "") -> None:
    ensure_outdir(out_dir)
    for t in tables:
        page = t.get("page", "unknown")
        on_page_idx = t.get("index_on_page", t.get("global_index", 0))
        name = f"{prefix}page_{page}_t{on_page_idx}.csv"
        path = os.path.join(out_dir, name)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in t.get("matrix", []):
                writer.writerow(row)

def export_merged_cells_csv(tables: List[Dict[str, Any]], out_path: str) -> None:
    ensure_outdir(os.path.dirname(out_path) or ".")
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "page", "table_global_index", "table_index_on_page",
            "anchor_row", "anchor_col", "row_span", "col_span",
            "text", "confidence", "covers"
        ])
        for t in tables:
            merged_cells = get_merged_cells_from_table(t)
            for m in merged_cells:
                covers_str = ";".join([f"{r}:{c}" for (r, c) in m.get("covers", [])])
                writer.writerow([
                    t.get("page"),
                    t.get("global_index"),
                    t.get("index_on_page"),
                    m.get("row_index"),
                    m.get("column_index"),
                    m.get("row_span", 1),
                    m.get("column_span", 1),
                    m.get("text", ""),
                    m.get("confidence", ""),
                    covers_str
                ])

def dump_tables_metadata_json(tables: List[Dict[str, Any]], out_path: str) -> None:
    ensure_outdir(os.path.dirname(out_path) or ".")
    # Everything is JSON-serializable already
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(tables, f, ensure_ascii=False, indent=2)

def save_tables_to_excel_with_merges(tables: List[Dict[str, Any]], out_xlsx: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for t in tables:
        page = t.get("page")
        on_page_idx = t.get("index_on_page", t.get("global_index", 0))
        sheet_name = f"p{page}_t{on_page_idx}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Write matrix
        matrix = t.get("matrix", [])
        for row in matrix:
            ws.append(row)

        # Apply merges
        for m in get_merged_cells_from_table(t):
            r1 = int(m["row_index"])
            c1 = int(m["column_index"])
            r2 = r1 + int(m.get("row_span", 1)) - 1
            c2 = c1 + int(m.get("column_span", 1)) - 1
            if r2 > r1 or c2 > c1:
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                cell = ws.cell(row=r1, column=c1)
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    ensure_outdir(os.path.dirname(out_xlsx) or ".")
    wb.save(out_xlsx)

# =========================
# CLI
# =========================

def parse_args():
    p = argparse.ArgumentParser(description="Extract tables and merged-cell info from AWS Textract JSON.")
    p.add_argument("--input", required=True, help="Path to Textract JSON file or directory of JSON files.")
    p.add_argument("--outdir", default="tables_out", help="Output directory.")
    p.add_argument("--prefix", default="", help="Filename prefix for CSV exports.")
    p.add_argument("--csv", action="store_true", help="Write per-table CSV files.")
    p.add_argument("--excel", action="store_true", help="Write one Excel workbook with real merged ranges.")
    p.add_argument("--merged-csv", action="store_true", help="Write a merged-cells summary CSV.")
    p.add_argument("--json", action="store_true", help="Write a JSON metadata dump for all tables.")
    p.add_argument("--header-rows", type=int, default=2, help="For heuristics when you care about merged headers only.")
    return p.parse_args()

def main():
    args = parse_args()

    data = load_textract_json_any(args.input)
    tables = extract_tables(data)

    if not tables:
        print("No TABLE blocks found.")
        return

    # Exports
    if args.csv:
        save_tables_to_csv(tables, args.outdir, prefix=args.prefix)

    if args.excel:
        xlsx_path = os.path.join(args.outdir, f"{args.prefix}textract_tables_with_merges.xlsx")
        save_tables_to_excel_with_merges(tables, xlsx_path)

    if args.merged_csv:
        merged_csv_path = os.path.join(args.outdir, f"{args.prefix}merged_cells_summary.csv")
        export_merged_cells_csv(tables, merged_csv_path)

    if args.json:
        json_path = os.path.join(args.outdir, f"{args.prefix}tables_metadata.json")
        dump_tables_metadata_json(tables, json_path)

    # Optional: show header-like merges in console
    header_counts = []
    for t in tables:
        header_merged = merged_header_cells_only(t, max_header_rows=args.header_rows)
        header_counts.append(len(header_merged))
    print(f"Extracted {len(tables)} tables across pages. Header-like merged cells per table: {header_counts}")

if __name__ == "__main__":
    main()


# Install

# pip install openpyxl
# Run examples

# Minimal (per-table CSVs only):
# python textract_tables_tool.py --input textract_output.json --outdir out --csv
# Everything (CSV, Excel with merges, merged-cells summary, JSON metadata):
# python textract_tables_tool.py --input textract_output.json --outdir out --csv --excel --merged-csv --json
# If you saved multiple Textract JSON files (e.g., async job pages) in a folder:
# python textract_tables_tool.py --input ./textract_json_shards --outdir out --csv --excel --merged-csv --json
# What outputs you’ll get

# out/page_{N}_t{K}.csv: One CSV per table (text only).
# out/textract_tables_with_merges.xlsx: One workbook with a sheet per table, real merged ranges applied.
# out/merged_cells_summary.csv: One row per merged cell with anchor row/col, row_span, col_span, text, confidence, and covered grid positions.
# out/tables_metadata.json: Full structured dump including per-cell metadata (row/col, spans, confidence, bbox/polygon).
# Notes and tips

# CSV does not preserve merged cells; use the merged summary CSV and/or Excel file to retain merge information.
# Cells are addressed with 1-based indices (Textract standard).
# Coordinates (bbox/polygon) are normalized 0–1 relative to page. Multiply by page width/height if you need absolute units.
# To focus only on merged headers, use merged_header_cells_only(table, max_header_rows=2) in custom logic or filter merged_cells_summary.csv by anchor_row <= 2.
# If you also deal with invoices/expenses, try Textract AnalyzeExpense; but for itemized medical bills, table parsing often gives you complete row structures.

# Got it. Below is a production-style solution to:

# Automatically detect header rows for each table CSV (even when headers span multiple rows).
# Return how many rows to skip per CSV for concatenation.
# Optionally use the last header row as column names during concat.
# Uses dataclasses, typing, and simple, explainable heuristics based on content patterns.
# Notes

# Since CSVs don’t preserve merges, we infer headers by analyzing the first few rows: headers tend to be text-heavy and/or contain known column keywords; data rows are number/currency/date heavy.
# If you still have the Textract JSON/tables metadata, you can get perfect header counts (e.g., use row_index <= 2). With CSV alone, heuristics work well for itemized bills.
# Code (save as detect_csv_headers.py)

# Requires: pip install pandas

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import List, Dict, Optional

import pandas as pd


@dataclass
class HeaderDetectConfig:
    max_scan_rows: int = 4                 # only look at top 4 rows
    numeric_ratio_threshold: float = 0.5   # >= this => data-like
    header_numeric_max: float = 0.4        # <= this + text/keywords => header-like
    min_fill_ratio: float = 0.3            # % of non-empty cells required to trust a header row
    drop_empty_top_rows: bool = True
    header_keywords: List[str] = field(default_factory=lambda: [
        "date", "service", "description", "code", "cpt", "hcpcs", "ndc",
        "rev", "revenue", "dx", "diagnosis", "modifier", "units", "qty",
        "quantity", "charge", "amount", "total", "allowed", "payment",
        "adjustment", "balance", "patient", "responsibility", "account",
        "claim", "provider", "pos", "rate", "price", "type", "category"
    ])


@dataclass
class HeaderDetection:
    header_row_indices: List[int]          # which rows (0-based) in top 4 are headers
    rows_to_skip: int                      # last header row index + 1, else 0
    debug: List[Dict] = field(default_factory=list)


NUMERIC_RE = re.compile(r"""
    ^\s*
    (?:\$)?                           # optional dollar
    -?                                # optional negative
    (?:
        \d{1,3}(?:,\d{3})*(?:\.\d+)?  # 1,234.56
        |
        \d+(?:\.\d+)?                 # 123 or 123.45
    )
    \s*%?\s*$                         # optional percent
""", re.X)

DATE_RE = re.compile(r"""
    ^\s*
    (?:
        \d{1,2}[-/]\d{1,2}[-/]\d{2,4}     # 01/31/2024
        |
        \d{4}[-/]\d{1,2}[-/]\d{1,2}       # 2024-01-31
    )
    \s*$
""", re.X)

CODE_RE = re.compile(r"^\s*(?:\d{5}|[A-Z]\d{4}|\d{3,4})\s*$", re.I)  # CPT/HCPCS/rev-like


def looks_numeric_like(s: str) -> bool:
    if not s:
        return False
    t = s.strip()
    if not t:
        return False
    return bool(NUMERIC_RE.match(t) or DATE_RE.match(t) or CODE_RE.match(t))


def contains_alpha(s: str) -> bool:
    return bool(re.search(r"[A-Za-z]", s or ""))


def contains_header_keyword(s: str, keywords: List[str]) -> bool:
    if not s:
        return False
    low = s.lower()
    return any(k in low for k in keywords)


def detect_header_rows_in_top4(csv_path: str, cfg: Optional[HeaderDetectConfig] = None) -> HeaderDetection:
    cfg = cfg or HeaderDetectConfig()

    # Read only first max_scan_rows rows, no type inference
    try:
        head_df = pd.read_csv(
            csv_path,
            header=None,
            nrows=cfg.max_scan_rows,
            dtype=str,
            keep_default_na=False,
            engine="python",
            on_bad_lines="skip",
        )
    except Exception as e:
        # If the CSV is empty or malformed, return no headers
        return HeaderDetection(header_row_indices=[], rows_to_skip=0, debug=[{"error": str(e)}])

    if head_df.empty:
        return HeaderDetection(header_row_indices=[], rows_to_skip=0, debug=[{"note": "empty file"}])

    # Optionally skip leading fully-empty rows
    start_idx = 0
    if cfg.drop_empty_top_rows:
        while start_idx < len(head_df):
            row_vals = head_df.iloc[start_idx].tolist()
            if all((str(x).strip() == "" for x in row_vals)):
                start_idx += 1
            else:
                break

    header_rows: List[int] = []
    first_data_row_idx: Optional[int] = None
    debug: List[Dict] = []

    # Precompute column count for fill ratio
    col_count = head_df.shape[1] if head_df.shape[1] > 0 else 1

    for i in range(start_idx, len(head_df)):
        row = head_df.iloc[i].tolist()
        stripped = [str(x).strip() for x in row]
        nonempty_vals = [x for x in stripped if x != ""]
        nonempty_count = len(nonempty_vals)
        fill_ratio = (nonempty_count / col_count) if col_count else 0.0
        numeric_like = sum(1 for x in nonempty_vals if looks_numeric_like(x))
        alpha_like = sum(1 for x in nonempty_vals if contains_alpha(x))
        keyword_hits = sum(1 for x in nonempty_vals if contains_header_keyword(x, cfg.header_keywords))
        total_for_ratio = nonempty_count if nonempty_count else 1
        numeric_ratio = numeric_like / total_for_ratio
        alpha_ratio = alpha_like / total_for_ratio

        is_empty = nonempty_count == 0
        is_data_like = (not is_empty) and (numeric_ratio >= cfg.numeric_ratio_threshold)
        is_header_like = (
            not is_empty
            and fill_ratio >= cfg.min_fill_ratio
            and numeric_ratio <= cfg.header_numeric_max
            and (alpha_ratio >= 0.5 or keyword_hits >= 1)
        )

        debug.append({
            "row_index": i,
            "values": row,
            "nonempty_count": nonempty_count,
            "fill_ratio": round(fill_ratio, 3),
            "numeric_like": numeric_like,
            "alpha_like": alpha_like,
            "keyword_hits": keyword_hits,
            "numeric_ratio": round(numeric_ratio, 3),
            "alpha_ratio": round(alpha_ratio, 3),
            "is_header_like": is_header_like,
            "is_data_like": is_data_like,
            "is_empty": is_empty,
        })

        # Stop scanning past the configured window (already limited by read_csv)
        # Mark the first data-like row; header rows must be before it
        if is_data_like and first_data_row_idx is None:
            first_data_row_idx = i

    # Decide which rows are headers:
    # - contiguous block of header-like rows from start_idx up to before first_data_row_idx
    for d in debug:
        i = d["row_index"]
        if i < start_idx:
            continue
        if first_data_row_idx is not None and i >= first_data_row_idx:
            break
        if d["is_header_like"]:
            header_rows.append(i)
        else:
            # If we haven't seen any header rows yet, keep looking;
            # once we've collected at least one header row, stop on first non-header row.
            if header_rows:
                break

    # Fallbacks:
    # If we found no header rows but encountered a data row: assume rows before first data row are headers if they are not empty and not ultra-numeric
    if not header_rows and first_data_row_idx is not None:
        provisional = []
        for d in debug:
            i = d["row_index"]
            if i < start_idx or i >= first_data_row_idx:
                continue
            if (not d["is_empty"]) and (d["numeric_ratio"] <= 0.4):
                provisional.append(i)
        header_rows = provisional

    # If still none, default to 1 header row if the very first non-empty row is not numeric-heavy
    if not header_rows:
        for d in debug:
            if d["row_index"] < start_idx:
                continue
            if not d["is_empty"] and d["numeric_ratio"] <= 0.4:
                header_rows = [d["row_index"]]
                break

    rows_to_skip = (max(header_rows) + 1) if header_rows else 0
    return HeaderDetection(header_row_indices=header_rows, rows_to_skip=rows_to_skip, debug=debug)


# -----------------------------
# Example usage
# -----------------------------

if __name__ == "__main__":
    # Example: run on all CSVs in a folder
    folder = "tables_out"  # change to your folder with table CSVs
    cfg = HeaderDetectConfig()

    for fname in sorted(os.listdir(folder)):
        if not fname.lower().endswith(".csv"):
            continue
        fpath = os.path.join(folder, fname)
        result = detect_header_rows_in_top4(fpath, cfg)
        print(f"{fname}: header rows={result.header_row_indices}, rows_to_skip={result.rows_to_skip}")

        # Example: read data skipping headers
        # df = pd.read_csv(fpath, skiprows=result.rows_to_skip, header=None, dtype=str, keep_default_na=False)
        # ... proceed to concat

# How to use it when concatenating

# First, detect headers per file to get rows_to_skip.
# Then read each CSV with skiprows=rows_to_skip and header=None.
# Align columns as needed and concatenate.
# Example concat pattern:

for each file:
det = detect_header_rows_in_top4(file)
df = pd.read_csv(file, skiprows=det.rows_to_skip, header=None, dtype=str, keep_default_na=False)
add df to list
pd.concat(list_of_df, ignore_index=True)
